import json
import random
import time
from datetime import datetime

from openpyxl import load_workbook, Workbook
from playwright.sync_api import sync_playwright


# ---------------------------------------------------
# SETTINGS
# ---------------------------------------------------

CONTACT_FILE = "contacts.xlsx"

today = datetime.now().strftime("%Y-%m-%d")

JSON_REPORT = f"whatsapp_report_{today}.json"
EXCEL_REPORT = f"whatsapp_report_{today}.xlsx"


# ---------------------------------------------------
# RANDOM DELAY
# ---------------------------------------------------

def random_delay():
    seconds = random.randint(2, 5)
    print(f"Waiting {seconds} seconds...")
    time.sleep(seconds)


# ---------------------------------------------------
# READ CONTACTS FROM EXCEL
# ---------------------------------------------------

def read_contacts():

    contacts = []

    workbook = load_workbook(CONTACT_FILE)
    sheet = workbook.active

    # Skip header row
    for row in sheet.iter_rows(min_row=2, values_only=True):

        name = row[0]
        phone = row[1]
        message = row[2]

        if name and phone:

            contacts.append({
                "name": str(name),
                "phone": str(phone),
                "message": str(message) if message else ""
            })

    workbook.close()

    return contacts


# ---------------------------------------------------
# SAVE JSON REPORT
# ---------------------------------------------------

def save_json_report(results):

    with open(JSON_REPORT, "w", encoding="utf-8") as file:

        json.dump(
            results,
            file,
            indent=4,
            ensure_ascii=False
        )

    print(f"\nJSON report saved: {JSON_REPORT}")


# ---------------------------------------------------
# SAVE EXCEL REPORT
# ---------------------------------------------------

def save_excel_report(results):

    workbook = Workbook()
    sheet = workbook.active

    sheet.title = "WhatsApp Report"

    # Header
    sheet.append([
        "Name",
        "Phone",
        "Message",
        "Status",
        "Error",
        "Last 3 Messages",
        "Screenshot"
    ])

    # Data
    for result in results:

        last_messages = " | ".join(
            result.get("last_3_messages", [])
        )

        sheet.append([
            result["name"],
            result["phone"],
            result["message"],
            result["status"],
            result["error"],
            last_messages,
            result["screenshot"]
        ])

    workbook.save(EXCEL_REPORT)

    print(f"Excel report saved: {EXCEL_REPORT}")


# ---------------------------------------------------
# MAIN PROGRAM
# ---------------------------------------------------

def main():

    print("Reading contacts...")

    contacts = read_contacts()

    print(f"Total contacts found: {len(contacts)}")

    results = []


    # Start Playwright
    with sync_playwright() as p:

        print("\nOpening WhatsApp Web...")

        browser = p.chromium.launch(
            headless=False
        )

        page = browser.new_page()

        page.goto("https://web.whatsapp.com")

        print("\nWhatsApp Web opened.")

        print("If QR code is displayed, scan it using your phone.")

        # Give user time to scan QR code
        page.wait_for_timeout(15000)

        print("Waiting for WhatsApp Web to load...")

        try:

            page.wait_for_selector(
                'div[contenteditable="true"]',
                timeout=60000
            )

            print("WhatsApp Web is ready!")

        except Exception:

            print("Could not detect WhatsApp Web.")

            browser.close()

            return


        # ---------------------------------------------------
        # PROCESS EACH CONTACT
        # ---------------------------------------------------

        for contact in contacts:

            name = contact["name"]
            phone = contact["phone"]
            template = contact["message"]

            print("\n--------------------------------")
            print(f"Processing: {name}")
            print(f"Phone: {phone}")
            print("--------------------------------")

            result = {
                "name": name,
                "phone": phone,
                "message": "",
                "status": "Failed",
                "error": "",
                "last_3_messages": [],
                "screenshot": ""
            }

            try:

                # -----------------------------------------
                # PERSONALIZE MESSAGE
                # -----------------------------------------

                if template:

                    message = template.replace(
                        "{name}",
                        name
                    )

                else:

                    message = f"Hello {name}, this is a daily message."

                result["message"] = message

                # -----------------------------------------
                # OPEN CHAT USING PHONE NUMBER
                # -----------------------------------------

                print("Opening contact...")

                chat_url = (
                    "https://web.whatsapp.com/send"
                    "?phone="
                    + phone.replace("+", "")
                    + "&text="
                    + message
                )

                page.goto(chat_url)

                random_delay()

                # -----------------------------------------
                # WAIT FOR CHAT
                # -----------------------------------------

                try:

                    page.wait_for_selector(
                        'div[contenteditable="true"]',
                        timeout=30000
                    )

                except Exception:

                    raise Exception(
                        "Contact/chat could not be opened."
                    )


                # -----------------------------------------
                # FIND MESSAGE BOX
                # -----------------------------------------

                message_box = page.locator(
                    'div[contenteditable="true"]'
                ).last

                message_box.wait_for(
                    state="visible",
                    timeout=30000
                )

                random_delay()


                # -----------------------------------------
                # SEND MESSAGE
                # -----------------------------------------

                print("Sending message...")

                message_box.click()

                page.keyboard.press("Enter")

                random_delay()

                print("Message sent.")


                # -----------------------------------------
                # SCREENSHOT
                # -----------------------------------------

                screenshot_name = (
                    f"sent_{today}_{name.replace(' ', '_')}.png"
                )

                page.screenshot(
                    path=screenshot_name,
                    full_page=True
                )

                result["screenshot"] = screenshot_name

                print(
                    f"Screenshot saved: {screenshot_name}"
                )


                # -----------------------------------------
                # GET LAST 3 MESSAGES
                # -----------------------------------------

                print("Reading last 3 messages...")

                messages = []

                try:

                    message_elements = page.locator(
                        'div.message-in span.selectable-text'
                    )

                    count = message_elements.count()

                    start = max(0, count - 3)

                    for i in range(start, count):

                        text = message_elements.nth(i).inner_text()

                        if text.strip():

                            messages.append(text.strip())

                except Exception as e:

                    print(
                        "Could not extract messages:",
                        e
                    )


                result["last_3_messages"] = messages[-3:]


                # -----------------------------------------
                # SUCCESS
                # -----------------------------------------

                result["status"] = "Sent"

                print(f"SUCCESS: Message sent to {name}")


            except Exception as e:

                result["status"] = "Failed"

                result["error"] = str(e)

                print(
                    f"ERROR sending to {name}: {e}"
                )


            # Add result to report
            results.append(result)

            # Wait before next contact
            random_delay()


        # ---------------------------------------------------
        # CLOSE BROWSER
        # ---------------------------------------------------

        print("\nAll contacts processed.")

        browser.close()


    # ---------------------------------------------------
    # SAVE REPORTS
    # ---------------------------------------------------

    save_json_report(results)

    save_excel_report(results)

    print("\n================================")
    print("Automation completed!")
    print("================================")


# ---------------------------------------------------
# RUN PROGRAM
# ---------------------------------------------------

if __name__ == "__main__":
    main()